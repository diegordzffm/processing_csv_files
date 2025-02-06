from openpyxl import load_workbook

def find_words_in_text(excel_file, text_file):
  """
  This function reads words from an Excel file and checks their presence in a text file.

  Args:
      excel_file (str): Path to the Excel file.
      text_file (str): Path to the text file.

  Returns:
      list: A list of words found in both the Excel file and the text file.
  """
  # Load the Excel workbook
  workbook = load_workbook(excel_file)
  sheet = workbook.active  # Assuming the first sheet contains the words

  # Initialize an empty list to store common words
  common_words = []

  # Read words from each row (assuming one word per cell)
  for row in sheet.iter_rows():
    for cell in row:
      if cell.value:  # Check if cell has a value
        word = cell.value.strip().lower()  # Convert to lowercase and remove whitespaces
        common_words.append(word) if word in open(text_file).read().lower() else None

  return common_words

# Example usage
excel_file = "words.xlsx"
text_file = "text.txt"
common_words = find_words_in_text(excel_file, text_file)

if common_words:
  print("Common words found:")
  for word in common_words:
    print(word)
else:
  print("No common words found.")
